import pandas as pd
import yfinance as yf
import subprocess
import os
import sys
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# 0. PREPARAÇÃO DO AMBIENTE (AUTO-INSTALAÇÃO) --- fundamentus e yfinance possuem conflitos de dependências, essa função tenta mitigar isso
def preparar_ambiente():
    os.chdir(os.path.expanduser("~")) # muda para a pasta home para evitar erros de permissão do banco de dados SQLite
    
    libs = ["yfinance", "pandas", "openpyxl", "requests-cache"]
    for lib in libs:
        try:
            __import__(lib)
        except ImportError:
            print(f"Instalando {lib}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", lib])
    
    try:
        __import__("fundamentus")
    except ImportError:
        print("Instalando fundamentus (modo compatibilidade)...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "fundamentus", "--no-deps"]) # --no-deps evita conflitos com o yfinance 

preparar_ambiente()

import fundamentus

# --- 1. MAPAS E FUNÇÕES DE ESTILO ---
traducao_setores = {
    "Financial Services": "Serviços Financeiros",
    "Basic Materials": "Materiais Básicos",
    "Utilities": "Utilidades Públicas",
    "Energy": "Energia",
    "Technology": "Tecnologia",
    "Industrials": "Industrial",
    "Consumer Cyclical": "Consumo Cíclico",
    "Consumer Defensive": "Consumo Não Cíclico",
    "Healthcare": "Saúde",
    "Real Estate": "Imobiliário",
    "Communication Services": "Comunicações",
    "N/A": "Não Identificado"
}

def formatar_aba(worksheet):
    """Aplica bordas, cores e ajuste de colunas."""
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    cor_header = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = borda_fina
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = cor_header

    for column in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column), default=0)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 4

# --- 2. COLETA DE DADOS ---
papeis_solidos = ["ITUB4", "VALE3", "WEGE3", "PETR4"]
setores_alvo = ["Serviços Financeiros", "Utilidades Públicas", "Energia", "Materiais Básicos"]

try:
    df_fund_geral = fundamentus.get_resultado() # obtém todas as empresas de uma vez
except Exception as e:
    print(f"Erro ao acessar Fundamentus: {e}")
    df_fund_geral = pd.DataFrame()

dados_coletados = []
for papel in papeis_solidos:
    try:
        acao = papel + ".SA"
        info = yf.Ticker(acao).info
        
        setor_pt = traducao_setores.get(info.get("sector", "N/A"), info.get("sector", "N/A"))
        
        # Busca indicadores no fundamentus
        pl = df_fund_geral.loc[papel, 'pl'] if papel in df_fund_geral.index else "N/A"
        dy = df_fund_geral.loc[papel, 'dy'] * 100 if papel in df_fund_geral.index else 0
        
        dados_coletados.append({
            "Nome": info.get("longName", "N/A"),
            "Papel": papel,
            "Setor": setor_pt,
            "P/L": pl,
            "DY (%)": round(dy, 2),
            "Preço Atual": info.get("currentPrice", "N/A")
        })
        print(f"   -> {papel}: OK")
    except Exception as e:
        print(f"   -> {papel}: Erro ({e})")

df_tabela = pd.DataFrame(dados_coletados).sort_values(by="Nome")
filtro_empresas = df_tabela[df_tabela["Setor"].isin(setores_alvo)].copy()

# --- 3. DADOS HISTÓRICOS ---
dfs_historico = []
if not filtro_empresas.empty:
    print(f"\n2. Baixando histórico dos papéis filtrados...")
    for papel in filtro_empresas['Papel']:
        df = yf.download(papel + ".SA", period="5y", interval="1d", progress=False)
        if df.empty: continue
        
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
            
        df = df.reset_index()
        df.insert(0, "Papel", papel)
        dfs_historico.append(df)

df_final_hist = pd.concat(dfs_historico, ignore_index=True) if dfs_historico else pd.DataFrame()
if not df_final_hist.empty:
    df_final_hist['Date'] = df_final_hist['Date'].dt.date
    df_final_hist = df_final_hist.rename(columns={"Date": "Data", "Close": "Fechamento"})

# --- 4. EXPORTAÇÃO ---
caminho_completo = os.path.join(os.path.expanduser("~"), "Downloads", "relatorio_integrado.xlsx")

try:
    with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
        df_tabela.to_excel(writer, sheet_name="Todas Empresas", index=False)
        filtro_empresas.to_excel(writer, sheet_name="Filtro Setor", index=False)
        if not df_final_hist.empty:
            df_final_hist.to_excel(writer, sheet_name="Dados Históricos", index=False)
        
        for sheet_name in writer.sheets:
            formatar_aba(writer.sheets[sheet_name])
            
    print(f"\nCONCLUÍDO! Arquivo salvo em DOWNLOADS")
except PermissionError:
    print("\nERRO: Feche o Excel antes de rodar o script.")
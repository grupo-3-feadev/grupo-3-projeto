# Bibliotecas Padrão do Sistema 
import os
import time
import warnings
from datetime import datetime

# Manipulação e Análise de Dados 
import numpy as np
import pandas as pd

# Coleta de Dados (Web Scraping e APIs) 
import requests
import yfinance as yf
from bs4 import BeautifulSoup

# Machine Learning e Estatística 
from sklearn.cluster import KMeans
from sklearn.preprocessing import RobustScaler
from sklearn.decomposition import PCA
from scipy.spatial import ConvexHull

# Visualização de Dados 
import seaborn as sns
import matplotlib.pyplot as plt

# Formatação e Exportação (Excel) 
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# ==============================================================================
# CONFIGURAÇÕES GERAIS
# ==============================================================================

# Bloqueia avisos técnicos, desativa alertas de cópia do Pandas e define o tema visual
warnings.simplefilter(action='ignore', category=FutureWarning)
pd.options.mode.chained_assignment = None
plt.style.use('seaborn-v0_8-whitegrid')

# Parâmetros de rede e filtros de mercado (B3)
HEADERS = {'User-agent': 'Mozilla/5.0'}
SETORES_ALVO = [1, 2, 3, 4, 10, 11, 14, 16, 17, 20, 21, 24, 27, 30, 31, 34, 37, 38, 40, 41]
COLS_NUMERICAS = ['Cotação', 'Div.Yield', 'P/L', 'P/VP', 'EV/EBITDA', 'ROE', 'Mrg. Líq.', 'Cresc. Rec.5a', 'Liq. Corr.', 'Dív.Brut/ Patrim.']


# ==============================================================================
# 1. PREPARAÇÃO DA BASE DE DADOS 
# ==============================================================================

def limpar_valor(val):
    # Converte strings financeiras para float para cálculos.
    if isinstance(val, str):
        val = val.replace('.', '').replace(',', '.').replace('%', '').strip()
        try: return float(val)
        except ValueError: return 0.0

    return val

def obter_mapa_setores():
    # Realiza Web-Scraping para mapear IDs numéricos aos nomes dos setores.
    try:
        r = requests.get("https://www.fundamentus.com.br/buscaavancada.php", headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.content, 'html.parser')
        select = soup.find('select', {'name': 'setor'})
        return {int(o['value']): o.text.strip() for o in select.find_all('option') if o['value'].isdigit()} if select else {}
    
    except: return {}

def obter_dados_setor(id_setor, nome_setor):
    # Busca dados fundamentalistas de um setor específico no Fundamentus.
    try:
        r = requests.get(f"https://www.fundamentus.com.br/resultado.php?setor={id_setor}", headers=HEADERS, timeout=10)
        df = pd.read_html(r.content, decimal=",", thousands=".")[0]
        if df.empty: return pd.DataFrame()

    except: return pd.DataFrame()

    df['ID'], df['SETOR'] = id_setor, nome_setor
    df.rename(columns={"Papel": "PAPÉIS"}, inplace=True)
    
    # Seleção e Limpeza de colunas numéricas
    cols = [c for c in ['ID', 'SETOR', 'PAPÉIS'] + COLS_NUMERICAS if c in df.columns]
    df = df[cols]

    for c in COLS_NUMERICAS:
        if c in df.columns: df[c] = df[c].apply(limpar_valor)
        
    df['Payout'] = (df['Div.Yield'] * df['P/L']).round(2) if 'Div.Yield' in df and 'P/L' in df else 0.0

    return df

def processar_historico_mercado(df_fund):
    # Utiliza yfinance para obter histórico de 5 anos e calcular DY Médio real.
    print("\nIniciando análise de mercado (Preços + Dividendos)...")
    dfs_hist, mapa_dy = [], {}
    papeis = df_fund['PAPÉIS'].unique()
    ano_atual = pd.Timestamp.now().year
    inicio_analise = pd.Timestamp(f"{ano_atual - 5}-01-01")

    for i, ticker in enumerate(papeis):
        print(f"[{i+1}/{len(papeis)}] {ticker}...", end='\r')
        try:
            acao = yf.Ticker(f"{ticker}.SA")
            hist = acao.history(period="5y", auto_adjust=True).reset_index()
            if hist.empty: continue
            
            hist['Date'] = pd.to_datetime(hist['Date']).dt.tz_localize(None)
            hist = hist[(hist['Date'] >= inicio_analise) & (hist['Date'].dt.year < ano_atual)]
            if len(hist) < 1200: continue

            # Cálculo do Dividend Yield Médio de 5 anos
            divs = acao.dividends.tz_localize(None) if not acao.dividends.empty else pd.Series(dtype=float)
            soma_divs = divs[divs.index >= inicio_analise].sum()
            preco_atual = df_fund.loc[df_fund['PAPÉIS'] == ticker, 'Cotação'].values[0]
            dy_medio = ((soma_divs / 5) / preco_atual * 100) if preco_atual > 0 else 0.0
            mapa_dy[ticker] = round(dy_medio, 2)

            # Formatação para o relatório
            hist.rename(columns={"Date": "Data", "Close": "Fechamento", "Volume": "Volume", "Open": "Abertura", "High": "Alta", "Low": "Baixa"}, inplace=True)
            hist['Data'] = hist['Data'].dt.date
            hist.insert(0, "Papel", f"{ticker}.SA")
            dfs_hist.append(hist[['Data', 'Papel', 'Abertura', 'Alta', 'Baixa', 'Fechamento', 'Volume']])

        except: continue
    
    print(f"\nColeta finalizada.{' '*20}")
    df_fund['DY (Média 5a)'] = df_fund['PAPÉIS'].map(mapa_dy).fillna(0.0)

    return df_fund, (pd.concat(dfs_hist, ignore_index=True) if dfs_hist else pd.DataFrame())

def calcular_volatilidade(df_cotacoes):
    # Calcula a volatilidade anualizada média das ações.
    if df_cotacoes.empty: return pd.DataFrame()
    df = df_cotacoes.copy()
    df['Data'] = pd.to_datetime(df['Data'])
    df['Retorno'] = df.groupby('Papel')['Fechamento'].pct_change()
    
    vol = df.groupby(['Papel', df['Data'].dt.year])['Retorno'].std() * np.sqrt(252)
    vol_media = vol.groupby('Papel').mean().mul(100).round(2).reset_index(name='Volatilidade_Media')
    vol_media['PAPÉIS'] = vol_media['Papel'].str.replace('.SA', '', regex=False)
    
    vol_pivot = vol.unstack().mul(100).round(2).fillna('-')
    vol_pivot['Volatilidade_Media'] = vol_media.set_index('Papel')['Volatilidade_Media']

    return vol_pivot.reset_index(), vol_media

# ==============================================================================
# 2. MOTOR DE CLUSTERIZAÇÃO E RANKING (MACHINE LEARNING)
# ==============================================================================

def preparar_dados_cluster(df_fund, df_vol_media):
    # Trata outliers, aplica filtros de qualidade e escala os dados 
    df_full = pd.merge(df_fund, df_vol_media[['PAPÉIS', 'Volatilidade_Media']], on='PAPÉIS', how='inner')
    
    # Filtros financeiros
    mask = (df_full['Liq. Corr.'] > 0) & (df_full['EV/EBITDA'] > 0) & (df_full['Cotação'] > 3)
    if 'P/L' in df_full: mask &= df_full['P/L'].between(-50, 100)
    if 'ROE' in df_full: mask &= df_full['ROE'].between(-50, 100)
    
    df_full = df_full[mask].sort_values('DY (Média 5a)', ascending=False)
    
    # Preparação para K-Means
    cols_ml = ['DY (Média 5a)', 'P/L', 'ROE', 'Volatilidade_Media']
    df_model = df_full[['PAPÉIS'] + cols_ml].dropna()
    df_model = df_model[(df_model['Volatilidade_Media'] < 100) & np.isfinite(df_model[cols_ml]).all(axis=1)]

    # Clip Outliers 
    for c in cols_ml:
        df_model[c] = df_model[c].clip(df_model[c].quantile(0.01), df_model[c].quantile(0.99))
        
    return df_full, df_model, RobustScaler().fit_transform(df_model[cols_ml])

def calcular_ranking(df):
    # Gera SCORE baseado em múltiplos critérios (Dívida, Crescimento, Liquidez).
    df_r = df.copy()
    criterios = {'Dív.Brut/ Patrim.': True, 'Cresc. Rec.5a': False, 'Liq. Corr.': False, 'Payout': False}
    df_r['SCORE'] = 0
    for col, asc in criterios.items():
        if col in df_r.columns:
            df_r['SCORE'] += df_r.groupby('Cluster')[col].rank(ascending=asc)
    return df_r.sort_values(['Cluster', 'SCORE'])[['Cluster', 'PAPÉIS', 'DY (Média 5a)'] + list(criterios.keys())]

# ==============================================================================
# 3. VISUALIZAÇÃO E EXPORTAÇÃO
# ==============================================================================

def salvar_graficos(dados_scaled, labels, k_ideal, df_resumo, pasta):
    # Gera visualizações: Método do Cotovelo, Clusters PCA e Risco x Retorno.
    # 1. Gráfico do Cotovelo
    inercias = [KMeans(k, n_init=10, random_state=42).fit(dados_scaled).inertia_ for k in range(2, 12)]
    plt.figure(figsize=(10, 6))
    plt.plot(range(2, 12), inercias, 'o-', color='#1f77b4')
    plt.plot(k_ideal, inercias[k_ideal-2], 'ro', markersize=12, label=f'K Ideal ({k_ideal})')
    plt.grid(True, linestyle='--'); plt.title('Método do Cotovelo'); plt.legend()
    plt.savefig(os.path.join(pasta, "Grafico_1_Cotovelo.png")); plt.close()

    # 2. Mapa de Clusters (PCA + ConvexHull)
    pca = PCA(n_components=2).fit_transform(dados_scaled)
    plt.figure(figsize=(12, 8))
    cores = sns.color_palette("bright", k_ideal)

    for i in range(k_ideal):
        pts = pca[labels == i]
        plt.scatter(pts[:,0], pts[:,1], c=[cores[i]]*len(pts), label=f'G{i}', s=60, alpha=0.8)

        if len(pts) >= 3:
            plt.fill(pts[ConvexHull(pts).vertices,0], pts[ConvexHull(pts).vertices,1], color=cores[i], alpha=0.15)

    plt.title('Mapa dos Grupos (PCA)'); plt.legend(); plt.savefig(os.path.join(pasta, "Grafico_2_Clusters_Poligonos.png")); plt.close()

    # 3. Relação Risco x Retorno
    plt.figure(figsize=(14, 9))
    sns.scatterplot(data=df_resumo, x='Volatilidade_Media', y='DY (Média 5a)', hue='Cluster', style='Cluster', palette='bright', s=100)

    for _, r in df_resumo[(df_resumo['DY (Média 5a)'] > 10) | (df_resumo['Volatilidade_Media'] < 25)].iterrows():
        plt.text(r['Volatilidade_Media']+0.2, r['DY (Média 5a)'], r['PAPÉIS'], size=8, weight='bold')

    plt.axvline(25, color='gray', ls='--'); plt.axhline(6, color='green', ls='--')
    plt.title('Risco x Retorno'); plt.legend(); plt.savefig(os.path.join(pasta, "Grafico_3_Risco_Retorno.png")); plt.close()

def formatar_excel(writer, sheet_name):
    # Aplica identidade visual (Azul/Branco) e auto-ajuste no Excel.
    ws = writer.sheets[sheet_name]
    fill, font = PatternFill("solid", fgColor="1F4E78"), Font(color="FFFFFF", bold=True)
    border, align = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin')), Alignment('center', 'center')
    
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max((len(str(c.value)) for c in col if c.value), default=10) + 4
        col[0].fill, col[0].font = fill, font
        for cell in col: cell.border, cell.alignment = border, align


# ==============================================================================
# ORQUESTRADOR PRINCIPAL (MAIN)
# ==============================================================================
def main():
    # --- FASE 1: COLETA E CONSOLIDAÇÃO ---
    print("--- FASE 1: COLETA DE DADOS ---")
    mapa_setores = obter_mapa_setores()
    dfs = [obter_dados_setor(sid, mapa_setores.get(sid, f"Setor_{sid}")) for sid in SETORES_ALVO]
    df_geral = pd.concat(dfs, ignore_index=True).drop_duplicates(subset=['PAPÉIS'])
    df_geral = df_geral[df_geral['Liq. Corr.'] > 0]
    
    if df_geral.empty: return print("Nenhum dado encontrado.")

    df_fund_final, df_cotacoes = processar_historico_mercado(df_geral)
    if df_cotacoes.empty: return print("Falha no histórico.")
    
    df_vol_raw, df_vol_media = calcular_volatilidade(df_cotacoes)

    # --- FASE 2: ANÁLISE E AGRUPAMENTO ---
    print("\n--- FASE 2: ANÁLISE E AGRUPAMENTO ---")
    df_full, df_model, dados_scaled = preparar_dados_cluster(df_fund_final, df_vol_media)
    
    # Lógica Geométrica para o K-Ideal
    inercias = [KMeans(k, n_init=10, random_state=42).fit(dados_scaled).inertia_ for k in range(2, 12)]
    x1, y1, x2, y2 = 2, inercias[0], 11, inercias[-1]
    dists = [abs((y2-y1)*k - (x2-x1)*inercias[k-2] + x2*y1 - y2*x1)/np.sqrt((y2-y1)**2+(x2-x1)**2) for k in range(2, 12)]
    k_ideal = max(4, range(2, 12)[np.argmax(dists)])
    print(f">>> Grupos definidos: {k_ideal}")

    df_model['Cluster'] = KMeans(k_ideal, n_init=10, random_state=42).fit_predict(dados_scaled)
    df_resultado = pd.merge(df_full, df_model[['PAPÉIS', 'Cluster']], on='PAPÉIS')
    
    # --- FASE 3: RANKING E SAÍDA ---
    df_ranking = calcular_ranking(df_resultado)
    df_resumo = df_resultado.groupby('Cluster')[['DY (Média 5a)', 'P/L', 'ROE', 'Volatilidade_Media']].mean().reset_index()
    df_resumo['Qtd'] = df_resultado.groupby('Cluster').size().values

    pasta_out = os.path.join(os.path.expanduser("~"), "Downloads")
    salvar_graficos(dados_scaled, df_model['Cluster'], k_ideal, df_resultado, pasta_out)
    
    arquivo_xls = os.path.join(pasta_out, f"Relatorio{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    print(f"\nGerando Excel: {arquivo_xls}")
    
    with pd.ExcelWriter(arquivo_xls, engine='openpyxl') as writer:
        dict_abas = {
            'Ranking ': df_ranking,
            'Resumo Clusters': df_resumo,
            'Base Completa': df_resultado.sort_values(['Cluster', 'DY (Média 5a)'], ascending=[True, False]),
            'Setores': df_fund_final[['ID', 'SETOR', 'PAPÉIS']],
            'Dados Fundamentus': df_fund_final,
            'Histórico Preços': df_cotacoes,
            'Volatilidade': df_vol_raw
        }
        for nome, df in dict_abas.items():
            df.to_excel(writer, sheet_name=nome, index=False)
            formatar_excel(writer, nome)
            
    print("[SUCESSO] Processo concluído.")

if __name__ == "__main__":
    main()